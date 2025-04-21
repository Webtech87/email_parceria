import os
from google.auth.transport.requests import Request

from googleapiclient.discovery import build
# from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from google_auth_oauthlib.flow import InstalledAppFlow
from google.oauth2.credentials import Credentials
from googleapiclient.errors import HttpError


def save_to_excel(data):
    file_name = 'client_info.xlsx'

    if not os.path.exists(file_name):
        wb = Workbook()
        sheet = wb.active

        headers = ["Nome", "Apelido", "Telefone", "Email", "Parceiro(a)", "Procedimento"]
        bold_font = Font(bold=True)

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = bold_font

    else:
        wb = load_workbook(file_name)

    sheet = wb.active
    sheet.append(data)
    wb.save(file_name)


# Google Sheets logic
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

SAMPLE_SPREADSHEET_ID = "1Jr1gHIpyliXhP-2fE99hWEzoBUxZUUZAWVwMjI1X_4M"
SAMPLE_RANGE_NAME = "Folha1!A2:ZZ"


def main():
    """Shows basic usage of the Sheets API.
    Prints values from a sample spreadsheet.
    """
    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                "credentials.json", SCOPES
            )
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open("token.json", "w") as token:
            token.write(creds.to_json())

    try:
        service = build("sheets", "v4", credentials=creds)

        # Call the Sheets API
        sheet = service.spreadsheets()
        result = (
            sheet.values()
            .get(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=SAMPLE_RANGE_NAME)
            .execute()
        )
        values = result.get("values", [])

        if not values:
            print("No data found.")
            return

        print("Name, Major:")
        for row in values:
            # Print columns A and E, which correspond to indices 0 and 4.
            print(f"{row[0]}, {row[4]}")
    except HttpError as err:
        print(err)


'''def add_data_to_sheet(sheet_id, data, sheet_name='Folha1'):
    # Append data to the given Google Sheet.
    credentials = get_credentials()
    print('++++++++++++++++++++add_data_to_sheet',credentials)
    service = build('sheets', 'v4', credentials=credentials)

    sheets = service.spreadsheets().get(spreadsheetId=SAMPLE_SPREADSHEET_ID).execute()
    print('++++++++++++++++++++',sheets)

    # Prepare data to be added
    body = {'values': [data]}

    range_ = f'{sheet_name}'

    # Append data to the first sheet in the document
    service.spreadsheets().values().append(
        spreadsheetId=sheet_id,
        range=range_,  # Adjust this range if needed
        valueInputOption='RAW',  # Ensure rows are added, not overwritten
        insertDataOption='INSERT_ROWS',
        body=body
    ).execute()
'''
