# imports for working with excel
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# import necessary to implement Google Sheets
from .secret_CSRF import CLIENT_SECRET_FILE
from googleapiclient.discovery import build
from google.oauth2 import service_account

from flask_wtf import FlaskForm
from wtforms import StringField, SubmitField, SelectField, BooleanField, validators
from wtforms.validators import DataRequired, Regexp
from wtforms.fields import EmailField
from flask_babel import lazy_gettext as _

# Убедитесь, что partners_list и procedure_list определены глобально
partners_list = [
    "Selecione um parceiro(a)",
    "Claudia Vieira",
    "Elia Camões",
    "Márcia Monteiro Micropigmentação"
]

procedure_list = [
    "Selecione uma procedimento",
    "Microcirurgia cosmética Consulta",
    "Micro lifting de sobrancelha",
    "Micro blefaroplastia superior",
    "Micro blefaroplastia inferior",
    "Lifting do terço médio e inferior",
    "Micro cervicoplastia (papada e pescoço)",
    "Micro implante de sobrancelha",
    "Laser CO2"
]


class PartnerShipForm(FlaskForm):
    nome = StringField(
        label=_(u'Nome'),
        validators=[
            DataRequired(message=_(u"Este campo é obrigatório.")),
            validators.Length(min=4, max=25),
            Regexp(r'^[A-Za-zÀ-ÿ\s]+$', message=_(u"O nome pode conter apenas letras и espaços."))
        ],
        render_kw={"class": "form-control", "placeholder": _(u"Introduza o seu Nome")}
    )
    apelido = StringField(
        label=_(u'Apelido'),
        validators=[
            DataRequired(message=_(u"Este campo é obrigatório.")),
            validators.Length(min=3, max=25),
            Regexp(r'^[A-Za-zÀ-ÿ\s]+$', message=_(u"O nome pode conter apenas letras e espaços."))
        ],
        render_kw={"class": "form-control", "placeholder": _(u"Introduza o seu Apelido")}
    )
    tel = StringField(
        label=_(u'Telefone'),
        validators=[
            DataRequired(message=_(u"Este campo é obrigatório.")),
            validators.Length(min=9, max=14),
            Regexp(r'^\d+$', message=_("Apenas números são permitidos."))
        ],
        render_kw={
            "class": "form-control",
            "placeholder": _("Introduza o seu contacto telefónico"),
            "inputmode": "tel",  # Это помогает отображать цифровую клавиатуру на мобильном устройстве
            "pattern": "\\d*"  # Гарантирует прием только цифр
        }
    )
    email = EmailField(
        label=_(u'Email'),
        validators=[
            DataRequired(message=_(u"Este campo é obrigatório.")),
            validators.Email(),
            validators.Length(min=6, max=35)
        ],
        render_kw={"class": "form-control", "placeholder": _("Introduza o seu Email"), "type": "email"}
    )
    partner_ship_list = SelectField(
        label=_(u'Parceiro(a)'),
        choices=[("", _("Selecione um parceiro(a)"))] + [(partner, partner) for partner in partners_list],
        validators=[DataRequired(message=_("Por favor, selecione um parceiro(a)."))],
        render_kw={
            "class": "input_class_selection form-control",
            "placeholder": _("Selecione um parceiro(a)")
        }
    )
    procedure_list = SelectField(
        label=_(u'Procedimento'),
        choices=[("", _("Selecione uma procedimento"))] + [(procedure, procedure) for procedure in
                                                           procedure_list],
        validators=[DataRequired(message=_("Por favor, selecione um procedimento."))],
        render_kw={"class": "input_class_selection form-control"}
    )
    checkbox = BooleanField(
        label=_(
            u'Eu concordo em partilhar minhas informações de contato com o propósito de ser contactado pela Santiclinic'),
        default=False,
        validators=[DataRequired(message=_("Por favor, aceite os termos de responsabilidade."))],
        render_kw={"class": "custom_checkbox form-check-input"}
    )
    submit = SubmitField(
        label=_(u'Enviar'),
        render_kw={"class": "input_class_submit btn-success"}
    )


# Create or load Excel file and save information from the form
def save_to_excel(data):
    file_name = 'client_info.xlsx'

    if not os.path.exists(file_name):
        wb = Workbook()
        sheet = wb.active

        headers = ["Nome", "Apelido", "Telefone", "Email", "Parceiro(a)", "Procedimento"]
        bold_font = Font(bold=True)

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = bold_font

    else:
        wb = load_workbook(file_name)

    sheet = wb.active
    sheet.append(data)
    wb.save(file_name)


# Google Sheets logic
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']


# Get credentials
def get_credentials():
    credentials = service_account.Credentials.from_service_account_file(CLIENT_SECRET_FILE, scopes=SCOPES)

    return credentials


def search_spreadsheet_by_name(service, spreadsheet_name="clients_info"):
    # Search for a sheet by name in Google Drive.
    query = f"name = '{spreadsheet_name}' and mimeType = 'application/vnd.google-apps.spreadsheet'"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    spreadsheets = results.get('files', [])

    if spreadsheets:
        return spreadsheets[0]['id']  # Return the first sheet's ID found
    else:
        return None  # No sheet found


def share_sheet_with_email(sheet_id, email='felipe.piano@gmail.com'):
    # Share the sheet with a Google account
    credentials = service_account.Credentials.from_service_account_file(CLIENT_SECRET_FILE, scopes=SCOPES)
    drive_service = build('drive', 'v3', credentials=credentials)

    permission = {
        'type': 'user',
        'role': 'writer',  # Use 'writer' for edit access, 'reader' for view access
        'emailAddress': email
    }

    # Apply the permission (share the sheet)
    drive_service.permissions().create(
        fileId=sheet_id,
        body=permission,
        sendNotificationEmail=False  # Change to True if you want an email notification
    ).execute()

    print(f'Sheet shared with {email}')


# Create a new sheet if it doesn't exist, or return the existing sheet's ID.
def create_or_get_sheet(sheet_name='Sheet1'):
    credentials = get_credentials()

    # First, use the Drive API to check for an existing sheet
    drive_service = build('drive', 'v3', credentials=credentials)
    sheet_id = search_spreadsheet_by_name(drive_service, 'clients_info_new')

    if not sheet_id:
        # If no sheet exists, create a new one using the Sheets API
        sheets_service = build('sheets', 'v4', credentials=credentials)
        spreadsheet = {
            'properties': {'title': 'clients_info_new'}
        }
        sheet = sheets_service.spreadsheets().create(body=spreadsheet, fields='spreadsheetId').execute()
        sheet_id = sheet['spreadsheetId']

        # Define the headers to insert into the first row
        headers = ["Nome", "Apelido", "Telefone", "Email", "Parceiro(a)", "Procedimento"]

        # Prepare the request body to update the first row with headers
        body = {'values': [headers]}

        # Update the first row with headers
        range_ = f'{sheet_name}!A1:F1'
        sheets_service.spreadsheets().values().update(
            spreadsheetId=sheet_id,
            range=range_,
            valueInputOption='RAW',
            # Use 'RAW' if you just want the raw data inserted, 'USER_ENTERED' for Google Sheets functions
            body=body
        ).execute()

        share_sheet_with_email(sheet_id)

        # Wait for the sheet to be fully created and populated
        print(f"Headers inserted into sheet: {sheet_name}")

    else:
        print(f"Sheet '{sheet_name}' already exists with ID: {sheet_id}")
        share_sheet_with_email(sheet_id)

    return sheet_id


def add_data_to_sheet(sheet_id, data, sheet_name='Sheet1'):
    # Append data to the given Google Sheet.
    credentials = get_credentials()
    service = build('sheets', 'v4', credentials=credentials)

    # Prepare data to be added
    body = {'values': [data]}

    range_ = f'{sheet_name}'

    # Append data to the first sheet in the document
    service.spreadsheets().values().append(
        spreadsheetId=sheet_id,
        range=range_,  # Adjust this range if needed
        valueInputOption='RAW',  # Ensure rows are added, not overwritten
        insertDataOption='INSERT_ROWS',
        body=body
    ).execute()
